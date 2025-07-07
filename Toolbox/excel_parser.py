import json
import os
import logging

from Tools.scripts.make_ctype import method
from openpyxl import load_workbook
from dataclasses import dataclass
from typing import Dict, List
from Toolbox.log_module import Logger

@dataclass
class APIDefinition:
    module_name: str
    api_name: str
    method: str
    endpoint: str
    headers: Dict[str, str]
    body_template: Dict[str, str]
    error_handling: List[str]

class ExcelParser:
    """
    Excel驱动接口定义解析器
    """
    def __init__(self, excel_path: str):
        self.logger = Logger(name="ExcelParser").get_logger()
        self.wb = load_workbook(excel_path)
        self.sheet = self.wb.active
        
        # 表格结构校验
        required_headers = ['模块名', '接口名', '方法', '端点', '请求头', '请求体模板', '错误处理']
        for col, header in enumerate(required_headers, 1):
            if self.sheet.cell(row=1, column=col).value != header:
                raise ValueError(f"Excel模板缺少必要列头'{header}'（第{col}列）")

        # 检查测试用例相关字段
        test_case_headers = ['测试用例名', '预期结果', '参数组合']
        for col in range(8, 11):
            if not self.sheet.cell(row=1, column=col).value:
                self.logger.warning(f"第{col}列缺少测试用例相关字段，测试生成功能可能受限")

    def _parse_headers(self, row: int) -> Dict[str, str]:
        return {
            cell.value: self.sheet.cell(row=row, column=cell.column+1).value
            for cell in self.sheet[2] 
            if cell.value
        }

    def generate_code(self, output_dir: str) -> str:
        """
        生成符合PEP8规范的Python接口代码
        """
        code_template = """
import requests
from typing import Dict, Any
from Toolbox.log_module import Logger

class {class_name}:
    def __init__(self):
        self.logger = Logger(name="{module_name}").get_logger()
        self.base_url = "{base_url}"  # 从配置读取
    
    def {method_name}(self, params: Dict[str, Any]) -> requests.Response:
        '''调用{{api_name}}接口'''
        try:
            response = requests.{{method}}(
                url=self.base_url + "{{endpoint}}",
                headers={{headers}},
                {{body_expression}}
            )
            response.raise_for_status()
            return response
        {{error_handling}}
        """
        
        # 实现具体转换逻辑
        api_definitions = []
        for row in range(2, self.sheet.max_row + 1):
            try:
                definition = APIDefinition(
                    module_name=self.sheet.cell(row=row, column=1).value,
                    api_name=self.sheet.cell(row=row, column=2).value,
                    method=self.sheet.cell(row=row, column=3).value.upper(),
                    endpoint=self.sheet.cell(row=row, column=4).value,
                    headers=self._parse_headers(row),
                    body_template=json.loads(self.sheet.cell(row=row, column=6).value),
                    error_handling=self.sheet.cell(row=row, column=7).value.split(';')
                )
                api_definitions.append(definition)
            except Exception as e:
                self.logger.error(f"第{row}行数据解析失败: {str(e)}")
                raise

        # 异常处理模板
        error_template = """
        except Exception as e:
            self.logger.error(f"调用{api_name}失败: {str(e)}")
            raise APICallError(f"{api_name}接口调用异常") from e
        """

        # 动态参数替换
        param_mapping = {
            '{headers}': 'headers=params.get("headers", {}),',
            '{body_expression}': 'json=params.get("body", {})' if method == 'post' else 'params=params'
        }

        return os.path.join(output_dir, "api_client.py")

        # 生成测试用例文件
        test_case_path = os.path.join(output_dir, "test_api_client.py")
        with open(test_case_path, 'w', encoding='utf-8') as f:
            f.write("""
import pytest
from {module} import {class_name}

class Test{class_name}:
    @pytest.mark.parametrize('test_data', load_test_data())
    def test_{method_name}(self, test_data):
        client = {class_name}()
        response = client.{method_name}(test_data['params'])
        assert response.status_code == test_data['expected_status']
            """.format(
                module=os.path.basename(output_file),
                class_name=class_name,
                method_name=method_name
            ))

if __name__ == '__main__':
    parser = ExcelParser('template.xlsx')
    parser.generate_code('./output')